import os
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font
from rapidfuzz import process, fuzz

# COM imports for method 2 (preserve formatting)
try:
    import win32com.client
    from win32com.client import constants
except Exception:
    win32com = None
    constants = None


# ------------------ SAFE SHEET READER ------------------
def read_sheet_openpyxl(fp, sheet_name):
    try:
        wb = load_workbook(fp, read_only=True, data_only=True)
    except Exception:
        return None

    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return None

    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = rows[1:]

    return pd.DataFrame(data, columns=headers)


# ------------------ MAIN APP ------------------
class SheetProcessorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Batch Rename Sheets & Consolidate Excel Files")
        self.geometry("900x540")
        self.resizable(False, False)

        self.master_path = tk.StringVar()
        self.folder_path = tk.StringVar()
        self.master_map = {}

        self._build_ui()

    # ------------------ UI ------------------
    def _build_ui(self):
        frm = tk.Frame(self, padx=10, pady=10)
        frm.pack(fill=tk.BOTH, expand=True)

        tk.Label(frm, text="Master Excel (.xlsx):").grid(row=0, column=0, sticky="w")
        tk.Entry(frm, textvariable=self.master_path, width=70).grid(row=0, column=1)
        tk.Button(frm, text="Browse", command=self.browse_master).grid(row=0, column=2)

        tk.Label(frm, text="Folder to process:").grid(row=1, column=0, sticky="w", pady=6)
        tk.Entry(frm, textvariable=self.folder_path, width=70).grid(row=1, column=1)
        tk.Button(frm, text="Browse", command=self.browse_folder).grid(row=1, column=2)

        tk.Button(frm, text="Rename Sheets", command=self.on_run).grid(row=2, column=0, columnspan=3, sticky="ew", pady=8)
        tk.Button(frm, text="Consolidate Sheets (Preserve Formatting - Windows/Excel only)", command=self.on_consolidate).grid(row=3, column=0, columnspan=3, sticky="ew")

        tk.Label(frm, text="Log:").grid(row=4, column=0, sticky="nw", pady=8)
        self.log = scrolledtext.ScrolledText(frm, width=115, height=22, state="disabled")
        self.log.grid(row=5, column=0, columnspan=3)

    # ------------------ HELPERS ------------------
    def log_msg(self, msg):
        self.log.configure(state="normal")
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.log.configure(state="disabled")

    def browse_master(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if p:
            self.master_path.set(p)
            self.load_master()

    def browse_folder(self):
        p = filedialog.askdirectory()
        if p:
            self.folder_path.set(p)

    def load_master(self):
        if not self.master_path.get():
            messagebox.showwarning("No master file", "Please choose a master Excel file first.")
            return

        self.master_map.clear()
        try:
            wb = load_workbook(self.master_path.get(), read_only=True)
        except Exception as e:
            messagebox.showerror("Error loading master", f"Failed to load master workbook: {e}")
            return

        for ws in wb.worksheets:
            # read first row as headers (openpyxl indexing)
            headers = [str(c.value) if c.value else "" for c in ws[1]]
            self.master_map[ws.title] = headers
        wb.close()
        self.log_msg(f"Master loaded. Sheets: {list(self.master_map.keys())}")

    # ------------------ RENAME SHEETS ------------------
    def on_run(self):
        threading.Thread(target=self.rename_sheets, daemon=True).start()

    def rename_sheets(self):
        # Auto-trim trailing blanks in source files before renaming
        self.log_msg("Auto-trimming trailing blank rows before renaming (if any)...")
        try:
            self.clean_blank_rows_bulk()
        except Exception as e:
            self.log_msg(f"Auto-trim failed: {e}")

        files = self._collect_files()

        for path in files:
            fname = os.path.basename(path)
            try:
                wb = load_workbook(path)
            except InvalidFileException:
                self.log_msg(f"Skipped invalid file: {fname}")
                continue
            except Exception as e:
                self.log_msg(f"Error opening {fname}: {e}")
                continue

            changed = False
            for ws in wb.worksheets:
                old = ws.title
                match, _, _ = process.extractOne(old, self.master_map.keys(), scorer=fuzz.token_sort_ratio)
                if match and match != old:
                    try:
                        ws.title = match
                        changed = True
                        self.log_msg(f"{fname}: '{old}' → '{match}'")
                    except Exception as e:
                        self.log_msg(f"Failed to rename sheet '{old}' in {fname}: {e}")

            if changed:
                try:
                    wb.save(path)
                    self.log_msg(f"Saved renamed sheets in {fname}")
                except Exception as e:
                    self.log_msg(f"Failed to save {fname}: {e}")

            wb.close()

        self.log_msg("Sheet renaming complete.")

    # ------------------ CONSOLIDATION (COM / preserve formatting) ------------------
    def on_consolidate(self):
        threading.Thread(target=self.consolidate_sheets_com, daemon=True).start()

    def consolidate_sheets_com(self):
        # Windows / Excel / pywin32 required
        if os.name != "nt" or win32com is None:
            messagebox.showerror(
                "Not available",
                "This consolidation method requires Windows with Excel installed and the pywin32 package."
            )
            self.log_msg("COM consolidation unavailable: not Windows or pywin32 not installed.")
            return

        files = self._collect_files()
        if not files:
            self.log_msg("No files found.")
            return

        out = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not out:
            return

        # Start COM Excel
        excel = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.ScreenUpdating = False
            excel.DisplayAlerts = False
            excel.EnableEvents = False
            # set manual calculation for speed if possible
            try:
                excel.Calculation = constants.xlCalculationManual
            except Exception:
                # fallback if constants not available for some reason
                pass

            # Create a new workbook and prepare it
            dest_wb = excel.Workbooks.Add()
            # Delete extra default sheets so we can add named sheets as needed (keep at least one)
            while dest_wb.Worksheets.Count > 1:
                try:
                    dest_wb.Worksheets(dest_wb.Worksheets.Count).Delete()
                except Exception:
                    break

            master_sheet_names = list(self.master_map.keys())
            for i, sheet_name in enumerate(master_sheet_names):
                # Create / name destination worksheet
                if i == 0:
                    dest_ws = dest_wb.Worksheets(1)
                    try:
                        dest_ws.Name = sheet_name[:31]
                    except Exception:
                        pass
                else:
                    dest_ws = dest_wb.Worksheets.Add(After=dest_wb.Worksheets(dest_wb.Worksheets.Count))
                    try:
                        dest_ws.Name = sheet_name[:31]
                    except Exception:
                        pass

                dest_row = 1  # current write row in destination
                # We'll record pasted blocks so we can add Filename as the final column afterwards.
                pasted_blocks = []  # list of tuples (start_row, data_rows_count, filename)

                # iterate each source file and copy used ranges
                for fp in files:
                    try:
                        # Open source workbook read-only
                        wb = excel.Workbooks.Open(os.path.abspath(fp), ReadOnly=True)
                    except Exception as e:
                        self.log_msg(f"Failed to open {os.path.basename(fp)} via COM: {e}")
                        continue

                    try:
                        # Attempt to get worksheet by name
                        try:
                            src_ws = wb.Worksheets(sheet_name)
                        except Exception:
                            # sheet not present
                            wb.Close(False)
                            continue

                        used = src_ws.UsedRange
                        if used is None:
                            wb.Close(False)
                            continue

                        rows_count = used.Rows.Count
                        cols_count = used.Columns.Count

                        rows_copied = 0
                        header_included = False

                        # First append: copy header + data
                        if dest_row == 1:
                            # copy whole used range (header + data) into column 1
                            try:
                                used.Copy(dest_ws.Cells(dest_row, 1))
                                rows_copied = rows_count
                                header_included = True
                            except Exception:
                                # fallback to copy + paste
                                try:
                                    used.Copy()
                                    dest_ws.Range(dest_ws.Cells(dest_row, 1)).Select()
                                    dest_ws.Paste()
                                    rows_copied = rows_count
                                    header_included = True
                                except Exception as e:
                                    self.log_msg(f"Copy error from {os.path.basename(fp)} sheet {sheet_name}: {e}")
                                    rows_copied = 0
                        else:
                            # subsequent files: skip header row in source (copy rows 2..end)
                            if rows_count >= 2:
                                try:
                                    src_range = src_ws.Range(src_ws.Cells(2, 1), src_ws.Cells(rows_count, cols_count))
                                    try:
                                        src_range.Copy(dest_ws.Cells(dest_row, 1))
                                    except Exception:
                                        src_range.Copy()
                                        dest_ws.Range(dest_ws.Cells(dest_row, 1)).Select()
                                        dest_ws.Paste()
                                    rows_copied = rows_count - 1
                                    header_included = False
                                except Exception as e:
                                    self.log_msg(f"Copy error (skip header) from {os.path.basename(fp)} sheet {sheet_name}: {e}")
                                    rows_copied = 0
                            else:
                                rows_copied = 0

                        # Record block for later filename filling.
                        # We should fill filenames only for data rows (not header). If header_included is True,
                        # the data rows start at dest_row + 1 and count is rows_copied - 1.
                        if rows_copied > 0:
                            fname = os.path.splitext(os.path.basename(fp))[0]
                            if header_included:
                                data_start = dest_row + 1
                                data_count = rows_copied - 1
                            else:
                                data_start = dest_row
                                data_count = rows_copied

                            if data_count > 0:
                                pasted_blocks.append((data_start, data_count, fname))

                            dest_row += rows_copied

                        self.log_msg(f"Copied from {os.path.basename(fp)} -> {sheet_name} (now {dest_row-1} rows)")
                    except Exception as e:
                        self.log_msg(f"Error processing {os.path.basename(fp)} sheet {sheet_name}: {e}")
                    finally:
                        try:
                            wb.Close(False)
                        except Exception:
                            pass

                # After all files copied for this sheet, determine final used columns and append Filename at the end
                try:
                    # Re-evaluate used range to find how many columns were pasted
                    final_used = dest_ws.UsedRange
                    if final_used is None:
                        final_cols = 0
                    else:
                        final_cols = final_used.Columns.Count
                except Exception:
                    # fallback: if UsedRange fails, assume at least 1 column
                    final_cols = 1

                filename_col = final_cols + 1
                # Write header for filename
                try:
                    dest_ws.Cells(1, filename_col).Value = "Filename"
                    # Match header style by copying format from first header cell
                    try:
                        dest_ws.Cells(1, 1).Copy()  # copy format from first header column
                        dest_ws.Cells(1, filename_col).PasteSpecial(constants.xlPasteFormats)
                    except Exception:
                        # fallback: at least make it bold if copy fails
                        dest_ws.Cells(1, filename_col).Font.Bold = True

                except Exception:
                    pass

                # Fill filename for each pasted block
                for start_row, count_rows, fname in pasted_blocks:
                    try:
                        top_cell = dest_ws.Cells(start_row, filename_col)
                        bottom_cell = dest_ws.Cells(start_row + count_rows - 1, filename_col)
                        dest_range = dest_ws.Range(top_cell, bottom_cell)
                        dest_range.Value = fname
                    except Exception:
                        # fallback to row-by-row
                        try:
                            for rr in range(start_row, start_row + count_rows):
                                try:
                                    dest_ws.Cells(rr, filename_col).Value = fname
                                except Exception:
                                    pass
                        except Exception:
                            pass

                # Try to autofit columns on the destination sheet (best-effort)
                try:
                    dest_ws.Columns.AutoFit()
                except Exception:
                    pass

            # Save and close destination workbook
            try:
                out_abs = os.path.abspath(out)
                dest_wb.SaveAs(out_abs)
                dest_wb.Close()
                self.log_msg(f"Consolidation saved (with formatting): {out_abs}")
            except Exception as e:
                self.log_msg(f"Failed to save consolidated workbook: {e}")
                try:
                    dest_wb.Close(SaveChanges=False)
                except Exception:
                    pass

        except Exception as e:
            self.log_msg(f"Unexpected COM error: {e}")
            messagebox.showerror("COM error", f"An error occurred during COM consolidation:\n{e}")
        finally:
            # Restore Excel settings and quit
            if excel is not None:
                try:
                    excel.Calculation = constants.xlCalculationAutomatic
                except Exception:
                    pass
                try:
                    excel.ScreenUpdating = True
                    excel.DisplayAlerts = True
                    excel.EnableEvents = True
                except Exception:
                    pass
                try:
                    excel.Quit()
                except Exception:
                    pass

    # ------------------ CLEAN TRAILING BLANK ROWS ------------------
    def clean_blank_rows_bulk(self):
        """Iterate collected files and trim trailing blank rows in-place."""
        files = self._collect_files()
        if not files:
            self.log_msg("No files found to trim.")
            return

        self.log_msg("Trimming trailing blank rows in source files...")
        for fp in files:
            fname = os.path.basename(fp)
            try:
                changed = self._clean_blank_rows_in_wb(fp)
                if changed:
                    self.log_msg(f"Trimmed trailing blank rows in {fname}")
                else:
                    self.log_msg(f"No trailing blanks found in {fname}")
            except Exception as e:
                self.log_msg(f"Error trimming {fname}: {e}")

        self.log_msg("Blank-row trimming complete.")

    def _clean_blank_rows_in_wb(self, fp):
        """
        Open workbook, and for each worksheet delete rows after the last row
        that contains any non-empty value. Returns True if workbook was changed.
        """
        try:
            wb = load_workbook(fp)
        except Exception as e:
            raise

        changed = False
        try:
            for ws in wb.worksheets:
                max_row = ws.max_row or 0
                max_col = ws.max_column or 0

                # If sheet is empty, nothing to do
                if max_row == 0 or max_col == 0:
                    continue

                # Find last row that contains any non-empty value
                last_nonblank = 0
                # iterate bottom-up for efficiency
                for r in range(max_row, 0, -1):
                    row_has_value = False
                    for c in range(1, max_col + 1):
                        cell = ws.cell(row=r, column=c)
                        val = cell.value
                        # treat None and empty/whitespace-only strings as blank
                        if val is not None:
                            if isinstance(val, str):
                                if val.strip() != "":
                                    row_has_value = True
                                    break
                            else:
                                # any non-string non-None value counts as data (numbers, formulas, dates)
                                row_has_value = True
                                break
                    if row_has_value:
                        last_nonblank = r
                        break

                # If trailing blank rows exist, delete them
                if last_nonblank < max_row:
                    delete_count = max_row - last_nonblank
                    try:
                        ws.delete_rows(last_nonblank + 1, delete_count)
                        changed = True
                    except Exception:
                        # if delete_rows fails for some reason, attempt row-by-row deletion
                        try:
                            for dr in range(last_nonblank + 1, max_row + 1):
                                try:
                                    ws.delete_rows(last_nonblank + 1, 1)
                                except Exception:
                                    pass
                            changed = True
                        except Exception:
                            pass

            if changed:
                # Save changes back to the same file
                wb.save(fp)
        finally:
            try:
                wb.close()
            except Exception:
                pass

        return changed

    # ------------------ FILE COLLECTION ------------------
    def _collect_files(self):
        files = []

        if self.folder_path.get():
            for f in os.listdir(self.folder_path.get()):
                if f.lower().endswith(".xlsx") and not f.startswith("~$"):
                    fp = os.path.join(self.folder_path.get(), f)
                    if fp != self.master_path.get():
                        files.append(fp)

        return files


# ------------------ RUN ------------------
if __name__ == "__main__":
    app = SheetProcessorApp()
    app.mainloop()